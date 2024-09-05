import datetime
from openpyxl import load_workbook
from urllib.parse import quote
from time import sleep
import pyautogui
import subprocess

def format_date(date_string):
  if isinstance(date_string, datetime.datetime):
      date_string = date_string.strftime('%Y-%m-%d') 
  try:
      return datetime.datetime.strptime(date_string, '%Y-%m-%d').date()
  except ValueError:
      print(f"data informada:'{date_string}' é inválida.")
      return None
    
def main():
  try:
      # Carrega o arquivo Excel do caminho especificado
      planilha = load_workbook(r"C:\Users\Fulano\Desktop\planilhas\controle.xlsx")
      sheet = planilha.active
  except FileNotFoundError:
       # Se o arquivo não for encontrado, imprime uma mensagem de erro e encerra o programa
      print("Error: O arquivo não foi encontrado. Cheque o caminho.")
      exit()

  current_date = datetime.datetime.now().date()

  for row in sheet.iter_rows(min_row=2, max_col=4):
    name, value, raw_date, phone_number = row[0].value, row[1].value, row[3].value, row[4].value

    date = format_date(raw_date)
    
    if date and date <= current_date:
      
      formatted_date = date.strftime('%d/%m/%Y')
      message = f"Olá {name} seu boleto venceu no dia {formatted_date}."  
      
       # Cria a URL do WhatsApp com o número de telefone e a mensagem
      text = f"start whatsapp://send?phone={quote(phone_number)}^&text={quote(message)}"
      try:
        # Executa o comando para abrir o WhatsApp Desktop e enviar a mensagem
        subprocess.Popen(["cmd", "/C", text], shell=True)
        
        sleep(6)

       # Extrai as coordenadas do botão de envio de mensagem
        send_btn = pyautogui.locateCenterOnScreen('enviar.png')
        pyautogui.click(send_btn[0], send_btn[1])

        sleep(5)
      except:    
        print(f'Não foi possível enviar: {name} , valor: R${value}, vencimento : {date}')
        with open('erros.csv','a',newline='',encoding='utf-8') as arquivo:
            arquivo.write(f'{name},{phone_number}, {value}')

if __name__ == "__main__":
    main()